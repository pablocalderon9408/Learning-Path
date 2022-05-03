# Openpyxl
from openpyxl import load_workbook
import math

def analyze_file():
    workbook = load_workbook(filename="./BD EDC 2022 Fase 1 para modificar.xlsx")
    worksheet = workbook['Sheet1']
    data = []

    for row in worksheet.iter_rows(min_row=2, max_col=1, max_row=1607):
        for cell in row:
            data.append(cell.value)

    print(data)








    # for row in worksheet.iter_rows(min_row=0):
    #     import ipdb ; ipdb.set_trace()
    #     row_data = {}
    #     username_list = []
    #     for cell in row:
    #         username_list.append(cell.value)
    #         import ipdb ; ipdb.set_trace()
    #         try:
    #             row_data[worksheet[f'{cell.column}1'].value] = cell.value
    #             # row_data[worksheet[f'{cell.column_letter}1'].value] = cell.value
    #         except Exception as e:
    #             print(">>>>> Exception raised: ", e)
    #             continue
    #     data.append(row_data)

    # count = 1
    # print("... Will start processing users")


if __name__ == '__main__':
    analyze_file()